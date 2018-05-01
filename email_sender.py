import csv
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import datetime as dt
import requests
from lxml import html
import smtplib
from email.message import EmailMessage
import time

your_email = input('Input your email: ')
your_password = input('Input your email password: ')
server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(your_email, your_password)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:51.0) Gecko/20100101 Firefox/51.0',
}

file = load_workbook('sample.xlsx').active
print(time.ctime())
X = True
while X:
	for row in file.iter_rows(row_offset=1):
		product_name = row[0].value
		url = row[1].value
		min_price_for_alert = row[2].value
		max_price_for_alert = row[3].value
		css_class_of_cur_price = row[4].value
		crawls_per_day = row[5].value
		try:
			email_addresses = row[6].value.split(',')
		except AttributeError:
			email_addresses = ''
		if email_addresses:
			print(url)
			r = requests.get(url, headers=headers)
			tree = html.fromstring(r.text)
			current_price = int(float(tree.xpath('//*[@*="{}"]/text()'.format(css_class_of_cur_price))[-1].strip().replace('$', '')))
			if current_price <= min_price_for_alert or current_price >= max_price_for_alert:
				for email_address in email_addresses:
					print(email_address)
					subject_message = "Price limit reached for {}. Current price is {}".format(product_name, current_price)
					message = "{} isimli ürün {}\'i adresinde {} TL'nin altında {} TL'ye satılıyor.".format(product_name, url, min_price_for_alert, current_price)
					msg = EmailMessage()
					msg.set_content(message)
					msg['Subject'] = subject_message
					msg['From'] = your_email
					msg['To'] = email_address
					server.send_message(msg)
					print("Message Sent to {}".format(email_address))
	print(time.ctime())
	time.sleep(43200)
