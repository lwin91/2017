# -*- coding: utf-8 -*-
import csv
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import os
from selenium.webdriver.common.proxy import Proxy
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from selenium.common.exceptions import NoSuchElementException
import phonenumbers
from openpyxl import load_workbook
from urllib.parse import urljoin
from lxml import html

chromedriver = "C:/chromedriver"
chromeOptions = webdriver.ChromeOptions()
prefs = {"profile.managed_default_content_settings.images":2}
chromeOptions.add_experimental_option("prefs",prefs)
driver = webdriver.Chrome(chromedriver, chrome_options=chromeOptions)


wb = Workbook()
ws = wb.active
ws.append(['Url', 'Number'])

# csvfile = open('Url list.csv', 'r', encoding='utf-8', errors='ignore')
# reader = csv.reader(csvfile)
# for num, item in enumerate(reader, 1):
# 	url = item[0].strip()

wb2 = load_workbook('qry_test.xlsx')
ws2 = wb2.active

a = open('logfile.txt', 'w', encoding='utf-8')

for num, row in enumerate(ws2['A'], 1):
	url = row.value
	print(num, url)
	a.write(url + '\n')
	if url.startswith('http'):
		url = url
	else:
		url = 'http://{}'.format(url)
	try:
		driver.get(url)
	except:
		continue
	try:
		text = driver.page_source
	except:
		continue


	number = ''
	mobile_numbers = ''
	for match in phonenumbers.PhoneNumberMatcher(text, "IT"):
		number = phonenumbers.format_number(match.number, phonenumbers.PhoneNumberFormat.E164)
		mobile_number = number.replace('+39', '').strip()

		if mobile_number.startswith('3') and 8 < len(mobile_number) < 12:
			mobile_number = mobile_number
			mobile_numbers += mobile_number + ', '
	mobile_numbers = ', '.join(filter(None, set(mobile_numbers.split(', '))))
	if len(mobile_numbers) > 1:
		data = [url, mobile_numbers]
		ws.append(data)
		print(num, data)
		a.write(', '.join(data))
		wb.save('phone_numbers.xlsx')
	else:
		tree = html.fromstring(driver.page_source)
		links = set(tree.xpath('//a/@href'))
		for link in links:

			if link:
				if link.startswith('http'):
					link = link
				elif link.startswith('/'):
					link = urljoin(url, link)

				if url.split('://')[-1] in link.split('://')[-1]:
					if url.split('.')[-1] != link.split('.')[-1].strip('/#'):
						print(link)
						try:
							driver.get(link)
						except:
							continue
						try:
							text = driver.page_source
						except:
							continue
						number = ''
						mobile_numbers = ''
						for match in phonenumbers.PhoneNumberMatcher(text, "IT"):
							number = phonenumbers.format_number(match.number, phonenumbers.PhoneNumberFormat.E164)
							mobile_number = number.replace('+39', '').strip()

							if mobile_number.startswith('3') and 8 < len(mobile_number) < 12:
								mobile_number = mobile_number
								mobile_numbers += mobile_number + ', '
						mobile_numbers = ', '.join(filter(None, set(mobile_numbers.split(', '))))
						if len(mobile_numbers) > 1:
							data = [url, mobile_numbers]
							ws.append(data)
							print(num, data)
							a.write(', '.join(data))
							wb.save('phone_numbers.xlsx')


			



			
					
					






